"""
オリジナル
2022/04/19
H. Uryu, 1521512@ed.tus.ac.jp(2023卒)

編集者
R. Kaneko 1519032@ed.tus.ac.jp(2025卒)
K. Tomiyoshi 1522529@ed.tus.ac.jp(2024卒)
M. Taniguchi 1521536@ed.tus.ac.jp(2023卒)

詳しくはGoogle Drive内の"更新情報"を参照
"""
#default設定
d_interval = 0.1#[s]
d_V_min = -0.8#[V]
d_V_max = 0.8#[s]
d_V_step = 0.1#[V]
d_loop = 1#回
d_folderpath = 'C:/Users/higuchi/Desktop/IVスイープ'
d_x_label = "Voltage [V]"
d_y_label = "Current [A]"
datas = {'Temp[℃]':[],
         'Temp[K]':[],
         'Resistance[Ω]':[]}

import matplotlib.pyplot as plt
import os
import pyvisa as visa
import re
import threading
import time
import tkinter as tk
import numpy as np 
import datetime
from decimal import Decimal
from tkinter import filedialog
from tkinter import ttk
from numpy import format_float_scientific as sci

rm = visa.ResourceManager(r'C:\WINDOWS\system32\visa64.dll')
#dev_=rm.list_resources()
dev = rm.open_resource('GPIB0::1::INSTR')
dev.timeout = 5000

print(dev.query('*IDN?'))

#送信コマンド
def write(command):
    dev.write(command)

#受信コマンド
def query(command):
    dev.query(command)

#フォルダ選択
def set_folder_func():
    dir = 'C:\\'
    folder_path = filedialog.askdirectory(initialdir = dir)
    textbox["folderpath"].delete(0, tk.END)
    textbox["folderpath"].insert(tk.END, folder_path)  

#グラフ
def graph(x_list, y_list, plot, scatter):
    def para(dic):
        return {f'{k1}.{k2}' : v for k1, d in dic.items() for k2, v in d.items()} 
    config = {
        "font" :{
            "family":"Times New Roman",
            "size":14
            },
        "xtick" :{
                "direction":"in",
                "top":True,
                "major.width":1.2,
                "labelsize":20.0
            },
        "ytick" :{
                "direction":"in",
                "right":True,
                "major.width":1.2,
                "labelsize":20.0
            },
        "axes" :{
            "linewidth":1.2,
            "labelpad":10
            },
        
        "figure" :{
            "dpi":150
                }
        }
    
    plt.rcParams.update(para(config))
    
    fig=plt.figure()
    ax=fig.add_subplot()

    if plot == True:
        ax.plot(x_list, y_list)
    if scatter == True:
        ax.scatter(x_list, y_list)
        
    ax.set_xlabel(d_x_label)
    ax.set_ylabel(d_y_label)
    plt.show()

#強制停止
def stop_func():
    global stop_flag
    stop_flag = True
    swrite("測定中断")

#時間指定で測定する場合
def setTemperature(startTemp, endTemp):
    i = startTemp
    while i < 40:
        datas["Temp[℃]"].append(i)
        datas['Temp[K]'].append(i+273)
        i = i + 2
    while i < 95:
        datas["Temp[℃]"].append(i)
        datas["Temp[K]"].append(i+273)
        i = i + 1
    while i >= endTemp:
        datas["Temp[℃]"].append(i)
        datas["Temp[K]"].append(i+273)
        i = i - 1

#IVスイープ
def measure(start, stop, step, sweep_times, interval): 
    global stop_flag

    write("SN"+str(start)+","+str(stop)+","+str(step))#SNstart,stop,step
    write("OPR")#出力

    for _ in range(sweep_times):
        if stop_flag == True:
            write("SWSP")#実行中のスイープを停止
            break
        
        write("*TRG")
        time.sleep(interval)
        
        A=dev.query("N?")
        A_ = float(A[3:-2])
        A_list.append(A_)
        
        V=dev.query("SOV?")  
        V_ = float(V[3:-2])
        V_list.append(V_)

        if value == False:
            if not A_ == 0:
                print(f"{V_:.6f} V\r\n{sci(A_, precision = 6, exp_digits = 2)} A\r\n{sci(V_/A_, precision = 6, exp_digits = 2)} Ω\r\n")
            else: 
                print(f"{V_:.6f} V\r\n{sci(A_, precision = 6, exp_digits = 2)} A\r\nError Ω\r\n")
        if plot == True or scatter == True:
            graph(V_list, A_list, plot, scatter)
        
    write("SBY")



#ファイル出力
# def output(filepath, x_list, y_list, extension_index):
#     # def output_txt():
#     #     with open(filepath, 'w') as data:
#     #         for x, y in zip(x_list, y_list):
#     #             data.write(f"{str(x)} {str(y)}\n")    
    
#     # def output_csv():
#     #     import csv
#     #     with open(filepath, 'w', newline="") as data:
#     #         writer = csv.writer(data)
#     #         for x, y in zip(x_list, y_list):
#     #             writer.writerow([x, y])
                
#     def output_xlsx():
#         from openpyxl import Workbook
#         from openpyxl import load_workbook
        
#         wb = Workbook()
#         wb.save(filepath)
#         wb = load_workbook(filepath)
#         ws = wb['Sheet']
#         ws = wb.active
        
#         for i, (x_val, y_val) in enumerate(zip(x_list, y_list), 1):
#             ws.cell(i, 1, x_val)
#             ws.cell(i, 2, y_val)
            
#         wb.save(filepath)
#         wb.close()    
        
#     # if extension_index == 0:
#     #     output_txt()          
#     # if extension_index == 1:
#     #     output_csv()
#     if extension_index == 2:
#         output_xlsx()

#抵抗値表示(I-V特性表示)
def resistance(x_list, y_list, temperature, determinedTemp):
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    
    x = np.array([x_list, np.ones(len(x_list))])
    x = x.T
    sigma, b = np.linalg.lstsq(x, y_list)[0] #aは抵抗値の逆数(電気伝導度)
    a = 1/sigma
    #温度データと抵抗データを追加
    if not determinedTemp:
        datas['Temp[℃]'].append(temperature)
    datas['Resistance[Ω]'].append(a)
    
    def para(dic):
        return {f'{k1}.{k2}' : v for k1, d in dic.items() for k2, v in d.items()} 
    config = {
        "font" :{
            "family":"Times New Roman",
            "size":14
            },
        "xtick" :{
                "direction":"in",
                "top":True,
                "major.width":1.2,
                "labelsize":20.0
            },
        "ytick" :{
                "direction":"in",
                "right":True,
                "major.width":1.2,
                "labelsize":20.0
            },
        "axes" :{
            "linewidth":1.2,
            "labelpad":10
            },
        
        "figure" :{
            "figsize":(16, 9),
            "dpi":60
                }
        }
    plt.rcParams.update(para(config))   

    fig = plt.figure()
    ax = fig.add_subplot()
    ax.plot(x_list, y_list, "ro")
    ax.plot(x, (a*x+b), "b")
    ax.set_xlabel("Voltage [V]")
    ax.set_ylabel("Current [A]")

    root1 = tk.Toplevel()
    root1.title("Resistance")
    root1.geometry("850x540")
    root1.resizable(False, False)
    root1.lift()
    
    canvas = FigureCanvasTkAgg(fig, master = root1)
    canvas.get_tk_widget().pack()
    
    def create_label(config):
        for var in config:
            label[var[0]] = tk.Label(root1, text = var[0], background = '#B0E0E6', font = ('Times New Roman', "20"))
            label[var[0]].place(x = var[1] + var[2]*var[5], y = var[3] + var[4]*var[5])
    label = {}
    label_list = [[f'Resistance: {sci(1/a, precision = 4, exp_digits = 4)} [Ω]'], [f'y = {sci(1/a, precision = 4, exp_digits = 4)}x + {sci(b, precision = 4, exp_digits = 4)}'],]
    label_params = [[25, 0, 10, 30],[400, 0, 10, 30]]
    label_config = [[tag_] + con + [i] for tag, con in zip(label_list, label_params) for i, tag_ in enumerate(tag)]
    create_label(label_config)

    print(f'Resistance: {sci(1/a, precision = 4, exp_digits = 4)} [Ω]', f'y = {sci(a, precision = 4, exp_digits = 4)}x + {sci(b, precision = 4, exp_digits = 4)}')

#エクセル書き出し
def writeEXCEL(filepath, x, y, header):
    from openpyxl import Workbook
    from openpyxl import load_workbook
    
    wb = Workbook()
    wb.save(filepath)
    wb = load_workbook(filepath)
    ws = wb['Sheet']
    ws = wb.active
    
    if header:
        ws.cell(1, 1, datas.keys()[0])
        ws.cell(1, 2, datas.keys()[1])
        startIndex = 2
    else:
        startIndex = 1

    for i, (x_val, y_val) in enumerate(zip(x, y), startIndex):
        ws.cell(i, 1, x_val)
        ws.cell(i, 2, y_val)
        
    wb.save(filepath)
    wb.close()
    
#IVスイープ一回走らせる
def run_func(interval, loop, V_min, V_max, V_step, temperature, folderpath, dataname, upTemp, determinedTemp):
    global V_list, A_list, stop_flag
    V_list, A_list =[], []
    stop_flag = False
    # interval = float(spinbox["interval"].get())#[s]
    # loop = float(spinbox["loop"].get())#ループ回数
    # V_min = Decimal(spinbox["V_min"].get())#[V]
    # V_max = Decimal(spinbox["V_max"].get())#[V]
    # V_step = Decimal(spinbox["V_step"].get())#[V]
    #V_start = 0#[V]

    extension_box_index = combobox["ext"].current()
    extension = combobox["ext"].get()
    sweepmode = combobox["sweepmode"].current()
    
    step_chk = abs(V_max-V_min)/V_step
    
    write("*RST")#初期化
    write("M1")#トリガーモード HOLD
    write("OH1")#ヘッダON
    write("VF")#電圧発生
    write("F2")#電流測定
    write("MD2")#DCスイープモード
    write("R0")#オートレンジ
    
    #エラーチェック
    # if not chk0 == True:
    # folderpath = textbox["folderpath"].get()
    # filename = textbox["filename"].get()
        
    #     if not os.path.exists(folderpath):
    #         swrite("無効なフォルダーパスです")
    #         return
    #     if filename == "" and extension == ".csv":
    #         swrite("ファイル名を入力して下さい")
    #         return
    
    if upTemp:
        upORdown = '_昇温_'
    else:
        upORdown = '_降温_'
    filename = dataname + upORdown + temperature + '℃_' + str(datetime.date.today())
    filename = re.sub(r'[\\/:*?"<>|]+-', '_', filename)
    filepath = folderpath +'/I-Vfiles/' + filename + extension
    
    if loop.is_integer() == False:
        swrite("ループ回数は整数値を設定して下さい")
        return
    else:
        loop = int(loop)
    
    if float(step_chk).is_integer() == False:
        swrite("ステップ数が整数になるように設定して下さい")
        return
    
    swrite(str(temperature) + "℃ 測定中")

    #測定実行
    sweeptimes1 = int(step_chk)+1
    #sweeptimes2 = int((V_max-V_start)/V_step)+1
    #sweeptimes3 = int((V_start-V_min)/V_step)+1
    
    # if sweepmode == 0:
    #     for i in range(loop): 
    #         swrite(f"測定中: ループ {i+1}/{loop}")
    #         measure(V_start, V_max, V_step, sweeptimes2, interval, chk1, chk2, chk4)        
    #         measure(V_max, V_min, V_step, sweeptimes1, interval, chk1, chk2, chk4)
    #         measure(V_min, V_start, V_step, sweeptimes3, interval, chk1, chk2, chk4)
    #     swrite("測定終了")
        
    # if sweepmode == 1:
    #     for i in range(loop):
    #         swrite(f"測定中: ループ {i+1}/{loop}")
    #         measure(V_min, V_max, V_step, sweeptimes1, interval, chk1, chk2, chk4)
    #         measure(V_max, V_min, V_step, sweeptimes1, interval, chk1, chk2, chk4)
    #     swrite("測定終了")
    
    if sweepmode == 2:
        measure(V_min, V_max, V_step, sweeptimes1, interval)
        swrite(str(temperature) + "℃ 測定終了")

    #ファイルに出力する場合
    writeEXCEL(filepath, V_list, A_list, True)

    #抵抗値の記録
    resistance(V_list, A_list, temperature, determinedTemp)

def run_measurement():

    #入力情報読み取り
    interval = float(spinbox["interval"].get())#[s]
    loop = float(spinbox["loop"].get())#ループ回数
    V_min = Decimal(spinbox["V_min"].get())#[V]
    V_max = Decimal(spinbox["V_max"].get())#[V]
    V_step = Decimal(spinbox["V_step"].get())#[V]
    startTemp = Decimal(spinbox["startTemp"].get())
    endTemp = Decimal(spinbox["endTemp"].get())

    chk = checkbutton['測定温度は決定している'].get()
    chk0 = checkbutton['ファイルに出力する'].get()
    chk1 = checkbutton['プロットを表示する'].get()
    chk2 = checkbutton['散布図を表示する'].get()
    chk3 = checkbutton['IV特性を表示する'].get()
    chk4 = checkbutton['測定値を表示しない'].get()

    folderpath = textbox["folderpath"].get()
    # filename = textbox["filename"].get()
    dataname = textbox["dataname"].get()
    nameOption = textbox["nameOption"].get()

    extension = combobox["ext"].get()
    
    #エラーチェック
    if chk0 == True:
        if not os.path.exists(folderpath):
            swrite("無効なフォルダーパスです")
            return
        if filename == "":
            swrite("ファイル名を入力して下さい")
            return
    os.makedirs("I-Vfiles", exist_ok=True)

    if chk:
        i = startTemp
        setTemperature(startTemp, endTemp)
        upTemp = True
        #各温度についてIVスイープ測定
        for i in datas["Temp[℃]"]:
            #昇温or降温の判定
            if (upTemp==True) and (datas["Temp[℃]"][i] < datas["Temp[℃]"][i-1]):
                upTemp = False
            run_func(interval, loop, V_min, V_max, V_step, i, folderpath, dataname, upTemp, True)
            #writeEXCEL(filepath, V_list, A_list)
            #resistance(V_list, A_list, datas["Temp[℃]"][i])
            #if chk3:

            #1分待機
            if (i < 40) and (len(datas["Resistance[Ω]"]) < 30):
                time.sleep()
            else:
                time.sleep(60)
    
    #ファイル名の設定
    if nameOption == "":
        filename = dataname + '_温度依存性_' + datetime.date.today()
    else:
        filename = dataname + '_温度依存性_' + datetime.date.today() + '_' + nameOption
    #ファイル名に禁止文字があるとき_に変更
    filename = re.sub(r'[\\/:*?"<>|]+-', '_', filename)


    filepath = folderpath +'/' + filename + extension
    writeEXCEL(filepath, datas["Temp[℃]"], datas["Resistance[Ω]"], True)

#メイン処理
def exc_run_func():
    try:
        t1 = threading.Thread(target = run_measurement)
        t1.start()

    except:
        swrite("予期せぬエラーです")

#ウィンドウ
root = tk.Tk()
root.title("I-V Sweep ver1.4")
root.geometry("430x300")#横×縦
root.resizable(False, False)#ウィンドウサイズをフリーズ
root.lift()#最前面に表示

#ラベル
def create_label(config):
            for var in config: 
                if var[5] == True:
                    label[var[0]] = tk.Label(text= var[0], background= '#B0E0E6')
                else:
                    label[var[0]] = tk.Label(text= var[0])
                label[var[0]].place(x=var[1] + var[2]*var[6], y= var[3] + var[4]*var[6])

label = {} 
label_list = [['保存先のフォルダ', 'データ名','追加入力項目'],
              ['V_min [V]', 'V_max [V]', 'V_step [V]', '遅延 [s]', 'ループ回数','開始温度','終了温度'],
              ['※折り返し無しの場合、無効'],
              ['ファイル形式'],
              ['測定モード'],]
#x = a+bx, y=c+dxを満たす[a, b, c, d] + background   
label_params = [[25, 0, 10, 30, True],
                [40, 0, 75, 25, False],
                [40, 0, 195, 0, False],
                [290, 0, 40, 0, True],
                [215, 0, 185, 0, True],]
label_config = [[tag_] + con + [i] for tag, con in zip(label_list, label_params) for i, tag_ in enumerate(tag)]
create_label(label_config)

#テキストボックス
def create_textbox(config):
    for key, var in config.items():
        textbox[key] = ttk.Entry(width= var[0])
        textbox[key].place(x= var[1], y= var[2])
        textbox[key].insert(0, var[3])
        
textbox = {}
textbox_config = {
    #{tag :[wid, x, y, init]}
    "folderpath" :[38, 120, 10, d_folderpath],
    "dataname" :[25, 120, 40, ""],
    "nameOption" :[25, 120, 70, ""]
    }  
create_textbox(textbox_config)

#スピンボックス
def create_spinbox(config):
    for i, (key, var) in enumerate(config.items()):
        spinbox[key] = ttk.Spinbox(
            root, 
            width = 7,
            format = '%3.1f',
            from_ = var[0],
            to = var[1],
            increment = var[2],
            )            
        spinbox[key].place(x= 125, y= 75 + 25*i)
        spinbox[key].insert(0, var[3])

spinbox = {}
spinbox_config = {
    #{tag :[min, max, step, init]}
    "V_min" :[-30.0, 30.0, 0.1, d_V_min],
    "V_max" :[-30.0, 30.0, 0.1, d_V_max],
    "V_step" :[-30.0, 30.0, 0.1, d_V_step],
    "interval" :[0.0, 10000.0, 0.1, d_interval],
    "loop":[1, 10000, 1, d_loop],
    "startTemp":[10, 95, 1, 22],
    "endTemp":[10, 95, 1, 40]
    }
create_spinbox(spinbox_config)

#チェックボタン
def create_checkbutton(config):
    for i, (key, var) in enumerate(config.items()):
        checkbutton[key] = tk.BooleanVar()
        checkbutton[key].set(var)
        chk = ttk.Checkbutton(
            root,
            variable = checkbutton[key],
            text = key
            )
        chk.place(x= 230, y= 75 + 20*i)

checkbutton = {}
checkbutton_config = {
    #[text :bln]
    '測定温度は決定している' :True,
    'ファイルに出力する' :True,
    'プロットを表示する' :False,
    '散布図を表示する' :False,
    'IV特性を表示する' :True,
    '測定値を表示する':False,
    }
        
create_checkbutton(checkbutton_config)

#ボタン
def create_button(config):
    for key, var in config.items():
        button[key] = ttk.Button(
            root,
            text = key,
            width = var[0],
            padding = [var[1], var[2]],
            command = var[5],
            )
        button[key].place(x= var[3], y= var[4])
        
button = {}
button_config = {
    #{tag :[wid, pad_EW, pad_NS, x, y, command]}
    "参照": [8, 0, 0, 360, 9, set_folder_func],
    "実行": [12, 0, 10, 125, 225, exc_run_func],
    "強制終了": [12, 0, 10, 225, 225, stop_func],
    }
create_button(button_config)

#プルダウンリスト
def create_combobox(config):
    for key, var in config.items():
        combobox[key] = ttk.Combobox(
            root,
            width = var[0],
            justify = "left", 
            state = "readonly",
            values = var[1],
            )
        combobox[key].place(x= var[2], y= var[3])
        combobox[key].current(var[4])
        
combobox = {}
combobox_config = {
    #tag :[wid, [values], x, y, init]
    "ext": [4, [".txt", ".csv", ".xlsx"], 360, 40, 2],
    "sweepmode": [18, ["双方向スイープ", "単方向スイープ", "単方向(折り返し無し)"], 275, 185, 2],        
    }
create_combobox(combobox_config)

statusbar = tk.Label(root, text = "", bd = 1, relief = tk.SUNKEN, anchor = tk.W)
statusbar.pack(side = tk.BOTTOM, fill = tk.X)
def swrite(text):
    statusbar["text"] = text

root.mainloop()
